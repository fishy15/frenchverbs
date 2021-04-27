from bs4 import BeautifulSoup
import requests
import re
import queue
import sys
import pprint
from openpyxl import Workbook
import datetime


class WordReference:

    def __init__(self, direction = "fren"):
        self.direction = direction
        self.url = "https://www.wordreference.com/" + direction

    def scrape(self, word):
        r = requests.get(self.url + '/' + word)
        sp = BeautifulSoup(r.content, "html.parser")
        translations = {}
        for table in sp.find_all("table", attrs = {"class" : "WRD"}):
            if str(table.get("id")) == "compound_forms":
                continue
            translation_html = table.find_all(attrs = {"class" : ["even", "odd"]})
            main_translation = None
            for translation in translation_html:
                to_word_tag = translation.find("td", attrs = {"class" : "ToWrd"})
                if to_word_tag is None:
                    continue

                if re.fullmatch(self.direction + r":\d+", translation.get("id", "")):
                    from_word_tag = translation.find("td", attrs = {"class" : "FrWrd"})
                    if from_word_tag is not None:
                        word_type = list(from_word_tag.find("em").strings)
                        if len(word_type) > 0 and word_type[0][0] != "v":
                            main_translation = None
                            continue
                        for unnecessary_tag in from_word_tag.strong.find_all("a"):
                            unnecessary_tag.decompose()
                        main_translation = re.sub(r"[;,(].*|/q\w\w", "",
                                from_word_tag.strong.text).strip()
                        if main_translation not in translations:
                            translations[main_translation] = []
                        translations[main_translation].append(set())

                if main_translation is not None and "(UK)":
                    possible_dialect = translation.find(attrs = { "class" : "To2" })
                    if possible_dialect and "(UK)" in possible_dialect.text:
                        continue
                    for unnecessary_tag in to_word_tag.find_all(["a", "em"]):
                        unnecessary_tag.decompose()
                    translations[main_translation][-1].update(
                            [re.sub(r"\s(?=\s|$)|/s\w{1,2}", "", synonym).strip() for synonym in
                            to_word_tag.text.split(",")])

        

        return translations

def parse_manual(answer):
    parens_matches = re.finditer(r"\(.*?\)", answer)
    split_inds = []
    for slash in list(re.finditer(r"/", answer))[::-1]:
        for parens in parens_matches:
            if slash.start() > parens.start() and slash.end() < parens.end():
                break
        else:
            split_inds.append(slash.start())
    split_inds.append(0)
    split_inds.reverse()
    split_inds.append(None)
    l = [answer[split_inds[i] + bool(i) : split_inds[i + 1]] for i in range(len(split_inds) - 1)]

    for i, synonym in enumerate(l):
        l[i] = re.sub(r"^to", "", re.sub(r"\([^/]*?\)", "", synonym).strip()).strip()
    return l

if __name__ == "__main__":
    wr = WordReference()
    with open(sys.argv[1], 'r', encoding='utf8') as f:
        manual_translations = dict([ translation.split("\t") for translation in f.readlines() ])
    with open(sys.argv[2], 'r', encoding='utf8') as f:
        verbs = f.readlines()
    automatic_translations = {}

    for i, verb in enumerate(verbs):
        verb = re.sub(r"\(.*?\)", "", verb).strip()
        verbs[i] = verb
        print(verb)
        automatic_translations.update(wr.scrape(verb))

    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint(automatic_translations)
    wb = Workbook()
    wb.create_sheet("Automated")
    wb.create_sheet("Manual")
    for k, v in automatic_translations.items():
        wb["Automated"].append([k] + [','.join(synonyms) for synonyms in v])
    for k, v in manual_translations.items():
        wb["Manual"].append([k, ','.join(parse_manual(v))])

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(datetime.datetime.now().strftime("%y%m%d_%H%M%S") + ".xlsx")
    #print(automatic_translations["changer"])

    #print(manual_translations["changer"])

